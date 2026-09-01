"""Bulk employee import from an .xlsx workbook.

Reuses EmployeeWriteSerializer for every row so imported employees go
through the exact same validation + user-provisioning + welcome-email
path as one created through the form — no second, divergent create path.
"""

from datetime import date, datetime

from openpyxl import load_workbook

from employees.models import Department, Designation
from employees.serializers import EmployeeWriteSerializer

# (template header, serializer field). Order defines the template columns.
IMPORT_COLUMNS = [
    ("First Name", "first_name"),
    ("Last Name", "last_name"),
    ("Email", "email"),
    ("Phone", "phone"),
    ("Date of Birth", "date_of_birth"),
    ("Gender", "gender"),
    ("Date Joined", "date_joined"),
    ("Employment Status", "employment_status"),
    ("Department", "department"),
    ("Designation", "designation"),
]

TEMPLATE_HEADERS = [h for h, _ in IMPORT_COLUMNS]
EXAMPLE_ROW = [
    "Sita", "Sharma", "sita.sharma@example.com", "+977-9800000000",
    "1995-04-12", "female", "2024-01-15", "active", "Engineering", "Software Engineer",
]
GENDER_CHOICES = ["male", "female", "other"]
STATUS_CHOICES = ["active", "on_leave", "resigned", "terminated"]


def _norm(value):
    return str(value).strip() if value is not None else ""


def _parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Accept common string forms.
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date '{value}' (use YYYY-MM-DD).")


def _find_header_row(ws):
    """Locate the header row by finding the one containing 'Email' — works
    whether the upload is our styled template (headers on row 4) or a
    plain sheet (row 1)."""
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        values = [_norm(c.value).lower() for c in ws[row_idx]]
        if "email" in values:
            return row_idx, values
    raise ValueError("Could not find a header row containing an 'Email' column.")


def preview_employees(file_obj):
    """Read the workbook and describe what *would* be imported. Creates nothing.

    Called by the import screen before anything is written, so somebody can see
    what a file contains and choose who to bring in.

    Each row comes back with a `status`:

    - `ready`      — will be created
    - `duplicate`  — that email is already an employee here
    - `invalid`    — no email, so there is nothing to create an account against

    Only `ready` rows are selectable; the others are shown rather than hidden,
    because "why is my row missing" is the question a silent filter produces.
    """
    from openpyxl import load_workbook

    from accounts.models import User

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    header_row, header_values = _find_header_row(ws)
    col_index = {}
    for header, field in IMPORT_COLUMNS:
        try:
            col_index[field] = header_values.index(header.lower())
        except ValueError:
            col_index[field] = None

    existing = {e.lower() for e in User.objects.values_list("email", flat=True) if e}
    seen_in_file = set()
    rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        def cell(field, row=row):
            idx = col_index.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        # Trailing blank rows Excel keeps around are not rows anybody typed.
        if not any(_norm(v) for v in row):
            continue

        email = _norm(cell("email")).lower()
        name = " ".join(filter(None, [_norm(cell("first_name")), _norm(cell("last_name"))]))

        if not email:
            status, note = "invalid", "No email — an account cannot be created without one."
        elif email in existing or email in seen_in_file:
            status, note = "duplicate", "Somebody with this email is already on the payroll here."
        else:
            status, note = "ready", ""
            seen_in_file.add(email)

        rows.append({
            "row": row_idx,
            "name": name or "—",
            "email": _norm(cell("email")),
            "department": _norm(cell("department")),
            "designation": _norm(cell("designation")),
            "status": status,
            "note": note,
        })

    return {"rows": rows}


def import_employees(file_obj, actor=None, rows=None):
    """Create one employee per data row. Returns {created, skipped,
    errors:[{row, email, error}]} — a bad row never aborts the rest.

    `rows`: the row numbers somebody chose in the preview. When given, every
    other row is left alone — which is the point of previewing at all.

    Omitted means "all of them", which keeps the original behaviour for any
    caller that has not been through a preview.
    """
    chosen = set(rows) if rows is not None else None
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    header_row, header_values = _find_header_row(ws)
    # Map our known headers → column index (0-based), by case-insensitive text.
    col_index = {}
    for header, field in IMPORT_COLUMNS:
        try:
            col_index[field] = header_values.index(header.lower())
        except ValueError:
            col_index[field] = None

    created, skipped, errors = 0, 0, []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # `row` is bound as a default so the closure captures *this*
        # iteration's row rather than whatever the loop variable ends on.
        def cell(field, row=row):
            idx = col_index.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        email = _norm(cell("email"))
        # Skip fully blank rows (trailing rows Excel keeps around).
        if not any(_norm(v) for v in row):
            continue
        # Not chosen in the preview — not an error, and not counted as skipped:
        # somebody deliberately left it out, and reporting it back as a problem
        # would make a considered choice look like a failure.
        if chosen is not None and row_idx not in chosen:
            continue
        if not email:
            skipped += 1
            errors.append({"row": row_idx, "email": "", "error": "Missing email — row skipped."})
            continue

        try:
            dept_name = _norm(cell("department"))
            desg_title = _norm(cell("designation"))
            department = Department.objects.get_or_create(name=dept_name)[0] if dept_name else None
            designation = None
            if desg_title:
                designation = Designation.objects.get_or_create(
                    title=desg_title, defaults={"department": department}
                )[0]

            data = {
                "first_name": _norm(cell("first_name")),
                "last_name": _norm(cell("last_name")),
                "email": email,
                "phone": _norm(cell("phone")),
                "gender": _norm(cell("gender")).lower(),
                "employment_status": _norm(cell("employment_status")).lower() or "active",
            }
            dob = _parse_date(cell("date_of_birth"))
            joined = _parse_date(cell("date_joined"))
            if dob:
                data["date_of_birth"] = dob
            data["date_joined"] = joined or date.today()
            if department:
                data["department"] = department.id
            if designation:
                data["designation"] = designation.id

            serializer = EmployeeWriteSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save(created_by=actor, updated_by=actor)
            created += 1
        except Exception as exc:  # noqa: BLE001 — surface any row error, keep going
            detail = getattr(exc, "detail", None)
            errors.append({"row": row_idx, "email": email, "error": _stringify_error(detail) or str(exc)})

    return {"created": created, "skipped": skipped, "errors": errors}


def _stringify_error(detail):
    if detail is None:
        return ""
    if isinstance(detail, dict):
        return "; ".join(f"{k}: {_stringify_error(v)}" for k, v in detail.items())
    if isinstance(detail, (list, tuple)):
        return ", ".join(_stringify_error(v) for v in detail)
    return str(detail)
