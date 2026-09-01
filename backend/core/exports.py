"""Styled .xlsx exports.

`xlsx_response` builds a branded workbook — title band, coloured header
row, zebra striping, borders, frozen header, auto-filter, auto-width — and
optionally highlights a status column by value and adds dropdown data
validation. `XlsxExportMixin` drops a filtered GET `/export/` onto any DRF
list viewset and dates the downloaded filename.
"""

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework.decorators import action

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

BRAND = "4F46E5"
_TITLE_FONT = Font(bold=True, size=15, color="111827")
_SUBTITLE_FONT = Font(size=10, color="6B7280")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor=BRAND)
_ZEBRA_FILL = PatternFill("solid", fgColor="F5F6FB")
_thin = Side(style="thin", color="E5E7EB")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center", vertical="center")

# value keyword -> fill; used to highlight a status/stage column.
_GREEN = PatternFill("solid", fgColor="D1FAE5")
_AMBER = PatternFill("solid", fgColor="FEF3C7")
_RED = PatternFill("solid", fgColor="FEE2E2")
_BLUE = PatternFill("solid", fgColor="DBEAFE")
_STATUS_FILLS = {
    # positive / done
    "active": _GREEN, "approved": _GREEN, "open": _GREEN, "completed": _GREEN,
    "hired": _GREEN, "present": _GREEN, "paid": _GREEN, "enrolled": _GREEN,
    # in-progress / attention
    "pending": _AMBER, "requested": _AMBER, "screening": _AMBER, "draft": _AMBER,
    "on leave": _AMBER, "late": _AMBER, "processing": _AMBER, "interview": _BLUE,
    "offer": _BLUE, "half day": _AMBER, "on_leave": _AMBER,
    # negative
    "rejected": _RED, "declined": _RED, "closed": _RED, "absent": _RED,
    "terminated": _RED, "no show": _RED, "no_show": _RED, "failed": _RED, "cancelled": _RED,
}


def _status_fill(value):
    return _STATUS_FILLS.get(str(value).strip().lower())


def xlsx_response(filename, headers, rows, title=None, subtitle=None, highlight_col=None, validations=None):
    """highlight_col: 1-based column index to colour by status value.
    validations: {1-based col index: [allowed strings]} for dropdown validation."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Export")[:31]
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    tcell = ws.cell(1, 1, title or "Export")
    tcell.font = _TITLE_FONT
    ws.merge_cells(f"A2:{last_col}2")
    ws.cell(2, 1, subtitle or f"Generated {timezone.localdate():%d %b %Y} · {len(rows)} rows").font = _SUBTITLE_FONT
    ws.row_dimensions[1].height = 22

    header_row = 4
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, i, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

    for r_offset, row in enumerate(rows):
        r = header_row + 1 + r_offset
        for i, value in enumerate(row, start=1):
            cell = ws.cell(r, i, value)
            cell.border = _BORDER
            if r_offset % 2 == 1:
                cell.fill = _ZEBRA_FILL
            if highlight_col == i:
                fill = _status_fill(value)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)
                    cell.alignment = _CENTER

    last_data_row = header_row + len(rows)
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{max(header_row, last_data_row)}"

    for col_idx, header in enumerate(headers, start=1):
        longest = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows if col_idx - 1 < len(r)] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(48, longest + 4)

    if validations and rows:
        for col_idx, allowed in validations.items():
            dv = DataValidation(type="list", formula1=f'"{",".join(allowed)}"', allowBlank=True)
            ws.add_data_validation(dv)
            col = get_column_letter(col_idx)
            dv.add(f"{col}{header_row + 1}:{col}{last_data_row}")

    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class XlsxExportMixin:
    """Adds a filtered GET `/export/` .xlsx action to a list viewset.

    Declare `export_headers` + `get_export_rows(queryset)`. Optionally set
    `export_highlight_header` (a header name whose cells get status-coloured)
    and `export_validations` ({header: [allowed values]} for dropdowns).
    The export respects the same filters/scoping as the list.
    """

    export_filename = "export.xlsx"
    export_title = "Export"
    export_headers: list[str] = []
    export_highlight_header: str | None = None
    export_validations: dict[str, list[str]] | None = None

    def get_export_rows(self, queryset):  # pragma: no cover - overridden
        raise NotImplementedError

    @action(detail=False, methods=["get"])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        rows = self.get_export_rows(queryset)

        highlight_col = None
        if self.export_highlight_header in self.export_headers:
            highlight_col = self.export_headers.index(self.export_highlight_header) + 1
        validations = None
        if self.export_validations:
            validations = {
                self.export_headers.index(h) + 1: vals
                for h, vals in self.export_validations.items()
                if h in self.export_headers
            }

        stem, _, ext = self.export_filename.rpartition(".")
        dated = f"{stem or self.export_filename}-{timezone.localdate():%Y-%m-%d}.{ext or 'xlsx'}"

        return xlsx_response(
            dated,
            self.export_headers,
            rows,
            title=self.export_title,
            highlight_col=highlight_col,
            validations=validations,
        )
